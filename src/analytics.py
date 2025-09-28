import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import pandas as pd
import matplotlib.pyplot as plt
import sqlalchemy
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from db_config import DB_URL




# подключение к БД
engine = sqlalchemy.create_engine(DB_URL)

def run_query(sql: str) -> pd.DataFrame:
    """Выполнить SQL-запрос и вернуть DataFrame"""
    return pd.read_sql(sql, engine)


# ---------------- ГРАФИКИ ---------------- #

def pie_players_by_team():
    sql = """
    SELECT t.Team AS team_name, COUNT(p.Player) AS num_players
    FROM players_stats p
    JOIN teams_ids t ON p.Teams = t.Team
    JOIN players_ids pi ON p.Player = pi.Player
    GROUP BY t.Team;
    """
    df = run_query(sql)
    df.set_index("team_name")["num_players"].plot.pie(autopct="%1.1f%%")
    plt.title("Распределение игроков по командам")
    plt.ylabel("")
    plt.savefig("charts/pie_players_by_team.png")
    plt.close()
    print(f"[OK] Pie chart saved ({len(df)} rows)")


def bar_avg_rating_by_team():
    sql = """
    SELECT t.Team AS team_name, ROUND(AVG(p.Rating), 2) AS avg_rating
    FROM players_stats p
    JOIN teams_ids t ON p.Teams = t.Team
    JOIN players_ids pi ON p.Player = pi.Player
    GROUP BY t.Team
    ORDER BY avg_rating DESC;
    """
    df = run_query(sql)
    df.plot(kind="bar", x="team_name", y="avg_rating", legend=False)
    plt.title("Средний рейтинг игроков по командам")
    plt.xlabel("Команды")
    plt.ylabel("Средний рейтинг")
    plt.tight_layout()
    plt.savefig("charts/bar_avg_rating_by_team.png")
    plt.close()
    print(f"[OK] Bar chart saved ({len(df)} rows)")


def hbar_top_kills():
    sql = """
    SELECT p.Player, SUM(p.Kills) AS total_kills, t.Team
    FROM players_stats p
    JOIN teams_ids t ON p.Teams = t.Team
    JOIN players_ids pi ON p.Player = pi.Player
    GROUP BY p.Player, t.Team
    ORDER BY total_kills DESC
    LIMIT 10;
    """
    df = run_query(sql)
    df.plot(kind="barh", x="Player", y="total_kills", legend=False)
    plt.title("Топ-10 игроков по убийствам")
    plt.xlabel("Убийства")
    plt.ylabel("Игрок")
    plt.tight_layout()
    plt.savefig("charts/hbar_top_kills.png")
    plt.close()
    print(f"[OK] Horizontal bar chart saved ({len(df)} rows)")


def line_maps_played():
    sql = """
    SELECT p.Tournament, t.Team, COUNT(p.`Rounds Played`) AS maps_played
    FROM players_stats p
    JOIN teams_ids t ON p.Teams = t.Team
    JOIN players_ids pi ON p.Player = pi.Player
    GROUP BY p.Tournament, t.Team
    ORDER BY p.Tournament;
    """
    df = run_query(sql)
    pivot = df.pivot(index="Tournament", columns="Team", values="maps_played").fillna(0)
    pivot.plot(kind="line", marker="o")
    plt.title("Карты, сыгранные командами по турнирам")
    plt.xlabel("Турнир")
    plt.ylabel("Количество карт")
    plt.legend(title="Команды", bbox_to_anchor=(1.05, 1), loc="upper left")
    plt.tight_layout()
    plt.savefig("charts/line_maps_played.png")
    plt.close()
    print(f"[OK] Line chart saved ({len(df)} rows)")


def bar_maps_played():
    sql = """
    SELECT Map, COUNT(*) AS times_played
    FROM maps_played
    GROUP BY Map
    ORDER BY times_played DESC;
    """
    df = run_query(sql)
    df.plot(kind="bar", x="Map", y="times_played", legend=False)
    plt.title("Популярность карт (по количеству игр)")
    plt.xlabel("Карта")
    plt.ylabel("Количество игр")
    plt.tight_layout()
    plt.savefig("charts/bar_maps_played.png")
    plt.close()
    print(f"[OK] Maps played chart saved ({len(df)} rows)")
    
def scatter_kd_vs_acs():
    sql = """
    SELECT p.Player,
           p.`Average Combat Score` AS acs,
           p.Kills,
           p.Deaths
    FROM players_stats p;
    """
    df = run_query(sql)

    # считаем K/D (делим на 1, чтобы избежать деления на ноль)
    df["kd_ratio"] = df["Kills"] / df["Deaths"].replace(0, 1)

    df.plot.scatter(x="acs", y="kd_ratio")
    plt.title("Зависимость K/D и Average Combat Score (ACS)")
    plt.xlabel("Average Combat Score")
    plt.ylabel("K/D ratio")
    plt.tight_layout()
    plt.savefig("charts/scatter_kd_vs_acs.png")
    plt.close()
    print(f"[OK] Scatter plot KD vs ACS saved ({len(df)} rows)")



def hist_multikills():
    sql = """
    SELECT Player,
           COALESCE(`2k`,0) + COALESCE(CAST(`3k` AS SIGNED),0) +
           COALESCE(CAST(`4k` AS SIGNED),0) + COALESCE(CAST(`5k` AS SIGNED),0) AS multikills
    FROM kills_stats;
    """
    df = run_query(sql)
    df["multikills"].plot.hist(bins=20)
    plt.title("Распределение мультикиллов у игроков")
    plt.xlabel("Количество мультикиллов")
    plt.ylabel("Частота игроков")
    plt.tight_layout()
    plt.savefig("charts/hist_multikills.png")
    plt.close()
    print(f"[OK] Histogram Multikills saved ({len(df)} rows)")



def bar_agents_pick_rate():
    sql = """
    SELECT Agent, AVG(CAST(`Pick Rate` AS DECIMAL(5,2))) AS avg_pick_rate
    FROM agents_pick_rates
    GROUP BY Agent
    ORDER BY avg_pick_rate DESC;
    """
    df = run_query(sql)
    df.plot(kind="bar", x="Agent", y="avg_pick_rate", legend=False)
    plt.title("Популярность агентов (средний Pick Rate)")
    plt.xlabel("Агент")
    plt.ylabel("Средний Pick Rate (%)")
    plt.tight_layout()
    plt.savefig("charts/bar_agents_pick_rate.png")
    plt.close()
    print(f"[OK] Agent pick rate chart saved ({len(df)} rows)")
    
    
def line_agents_by_stage():
    sql = """
    SELECT Stage, Agent, AVG(CAST(`Pick Rate` AS DECIMAL(5,2))) AS avg_pick_rate
    FROM agents_pick_rates
    GROUP BY Stage, Agent
    ORDER BY Stage;
    """
    df = run_query(sql)
    pivot = df.pivot(index="Stage", columns="Agent", values="avg_pick_rate").fillna(0)
    pivot.plot(kind="line", marker="o")
    plt.title("Популярность агентов по стадиям турниров")
    plt.xlabel("Стадия")
    plt.ylabel("Средний Pick Rate (%)")
    plt.legend(title="Агент", bbox_to_anchor=(1.05, 1), loc="upper left")
    plt.tight_layout()
    plt.savefig("charts/line_agents_by_stage.png")
    plt.close()
    print(f"[OK] Line chart saved ({len(df)} rows)")




# ---------------- EXCEL EXPORT ---------------- #

def export_to_excel(dfs: dict, filename: str):
    filepath = f"exports/{filename}"
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for sheet, df in dfs.items():
            df.to_excel(writer, sheet_name=sheet, index=False)

    wb = load_workbook(filepath)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.freeze_panes = "B2"

        # градиент для числовых колонок
        rule = ColorScaleRule(start_type="min", start_color="FFAA0000",
                              mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                              end_type="max", end_color="FF00AA00")
        max_row = ws.max_row
        max_col = ws.max_column
        ws.conditional_formatting.add(f"A2:{chr(64+max_col)}{max_row}", rule)

        ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)
    print(f"[OK] Created {filename}, {len(dfs)} sheets, {sum(len(df) for df in dfs.values())} rows")


# ---------------- PLOTLY TIME SLIDER ---------------- #
ROLES = {
    "Duelists":   ["phoenix", "jett", "reyna", "raze", "yoru", "neon", "iso", "waylay"],
    "Initiators": ["breach", "fade", "gekko", "kayo", "skye", "sova", "tejo"],
    "Sentinels":  ["chamber", "cypher", "deadlock", "killjoy", "sage", "vyse"],
    "Controllers":["astra", "brimstone", "clove", "harbor", "omen", "viper"]
}



def plotly_time_slider():
    sql = """
    SELECT Tournament, Stage, Agent,
           AVG(CAST(`Pick Rate` AS DECIMAL(6,2))) AS avg_pick_rate
    FROM agents_pick_rates
    GROUP BY Tournament, Stage, Agent
    ORDER BY Tournament, Stage;
    """
    df = run_query(sql)

    # выделим год из названия турнира
    df["Year"] = df["Tournament"].str.extract(r"(\d{4})")

    # строим scatter с ползунком
    fig = px.scatter(
        df,
        x="Stage",
        y="avg_pick_rate",
        color="Agent",
        animation_frame="Year",   # вот здесь появляется slider
        size="avg_pick_rate",
        hover_name="Agent",
        title="Популярность агентов по стадиям (слайдер по годам)"
    )

    fig.show()  # просто открывается в браузере

    


# ---------------- MAIN ---------------- #

if __name__ == "__main__":
    pie_players_by_team()
    bar_avg_rating_by_team()
    hbar_top_kills()
    line_maps_played()
    bar_maps_played()
    bar_agents_pick_rate()
    line_agents_by_stage()
    hist_multikills()
    scatter_kd_vs_acs()
    
    
    
    # пример экспорта
    df1 = run_query("SELECT * FROM players_stats LIMIT 100;")
    df2 = run_query("SELECT * FROM kills_stats LIMIT 100;")
    export_to_excel({"Players": df1, "Kills": df2}, "valorant_report.xlsx")

    # интерактивный график (показывается в браузере)
    plotly_time_slider()  # можно указать роль или None



